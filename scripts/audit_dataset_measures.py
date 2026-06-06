import csv
import io
import json
import os
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MAP_PATH = ROOT / "content" / "datahub-datasets.json"
ZIP_PATH = Path(r"C:\Users\sk3626\Downloads\Datahub-data-main (2).zip")
CSV_OUT = ROOT / "dataset_measure_inventory.csv"
MD_OUT = ROOT / "dataset_measure_audit.md"
MAX_SCAN_ROWS = 90
MAX_SCAN_COLS = 120

TOPICS = {
    "employment/labour market": ["employment", "labour", "neet", "skills", "vacancies", "occupation", "job", "work", "training", "earnings", "wage"],
    "housing": ["housing", "house", "affordability", "rent", "tenure", "stock", "dwelling", "land use", "developed land"],
    "business/industry": ["business", "industry", "sector", "enterprise", "investment", "firm", "gva", "productivity", "innovation", "trade", "export"],
    "health": ["health", "life expectancy", "mortality", "illness", "disease"],
    "environment/energy/emissions": ["greenhouse", "emissions", "co2", "carbon", "energy", "fuel", "environment", "sustainability"],
    "population": ["population", "migration", "age", "demographic"],
    "transport": ["transport", "travel", "commut", "access", "journey", "mode"],
    "GDP/GVA/productivity": ["gdp", "gva", "productivity", "gross domestic", "economic output"],
}

TOPIC_ORDER = list(TOPICS)
CATEGORY_TOPIC_HINTS = {
    "labour-market": ["employment/labour market"],
    "housing-and-land-use": ["housing"],
    "economy": ["business/industry", "GDP/GVA/productivity"],
    "health": ["health"],
    "environment": ["environment/energy/emissions"],
    "population": ["population"],
    "transport": ["transport"],
    "poverty-and-deprivation": ["poverty/deprivation"],
}
TOPIC_PREFERRED_CATEGORIES = {
    "employment/labour market": ["labour-market"],
    "housing": ["housing-and-land-use"],
    "business/industry": ["economy", "labour-market"],
    "health": ["health"],
    "environment/energy/emissions": ["environment"],
    "population": ["population"],
    "transport": ["transport"],
    "GDP/GVA/productivity": ["economy"],
}
TOPIC_SAMPLE_KEYWORDS = {
    "employment/labour market": ["employment", "labour", "neet", "skills", "training", "vacancies"],
    "housing": ["housing", "affordability", "tenure", "dwelling", "stock"],
    "business/industry": ["business", "industry", "enterprise", "sector", "trade", "investment"],
    "health": ["health", "life expectancy", "mortality"],
    "environment/energy/emissions": ["greenhouse", "emissions", "co2", "carbon", "energy"],
    "population": ["population", "migration", "births", "deaths"],
    "transport": ["transport", "travel", "commuting", "journey", "services"],
    "GDP/GVA/productivity": ["gdp", "gva", "productivity", "gross domestic", "economic output"],
}
SOURCE_DATA_SHEET_NAMES = [
    "raw",
    "raw data",
    "data",
    "source data",
    "source",
    "original data",
    "input data",
    "input",
    "lookup",
    "table",
    "tables",
    "detailed data",
    "underlying data",
    "observations",
    "values",
]
SOURCE_DATA_EXCLUDED_SHEET_NAMES = [
    "metadata",
    "meta",
    "notes",
    "contents",
    "readme",
    "methodology",
    "definitions",
    "analysis",
    "further analysis",
    "chart",
    "charts",
    "summary",
    "pivot",
    "dashboard",
]

BREAKDOWN_PATTERNS = {
    "local authority/geography": ["geography", "area", "local authority", "la name", "la code", "district", "region", "authority", "country", "place", "location"],
    "age": ["age", "age band", "age group"],
    "sex/gender": ["sex", "gender", "male", "female"],
    "sector/industry": ["sector", "industry", "sic", "industrial", "emitter", "emissions sector"],
    "occupation": ["occupation", "soc"],
    "qualification/skills": ["qualification", "skill", "skills", "training"],
    "tenure/housing type": ["tenure", "dwelling", "property", "housing type", "bedroom", "house type"],
    "fuel/energy type": ["fuel", "energy", "gas", "electricity", "co2", "greenhouse"],
    "transport mode/distance": ["mode", "distance", "travel", "commut", "journey", "transport"],
    "time/year": ["year", "date", "time", "quarter", "period", "month"],
    "business size": ["size band", "enterprise size", "employee size", "firm size", "employment size"],
    "health outcome": ["health", "life expectancy", "mortality", "death", "outcome"],
    "deprivation/income": ["income", "poverty", "deprivation", "imd"],
}

MEASURE_PATTERNS = {
    "rates/percentages/proportions": ["rate", "rates", "percent", "percentage", "proportion", "%"],
    "counts/numbers": ["count", "counts", "number", "numbers", "n_people", "headcount", "cohort"],
    "totals": ["total", "sum"],
    "numerators": ["numerator", "number neet", "number activity", "numerator count"],
    "denominators/base": ["denominator", "cohort", "base", "sample", "population base"],
    "confidence intervals": ["confidence", "ci", "lower", "upper", "lci", "uci", "interval"],
    "ratios/prices": ["ratio", "affordability", "price", "prices", "median", "mean"],
    "index values": ["index", "indices"],
    "currency/value": ["£", "gbp", "wage", "earnings", "income", "salary"],
    "emissions/energy totals": ["kt co2", "ktco2", "ktco2e", "co2", "emission", "energy", "kwh"],
    "time-series values": ["year", "date", "period"],
}


def norm(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u202f", " ").replace("\u00a0", " ").replace("\ufeff", "")
    text = text.replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def norm_path(value):
    return re.sub(r"/+", "/", norm(value).replace("\\", "/")).strip("/").lower()


def compact(values, limit=14):
    result = []
    seen = set()
    for value in values:
        text = norm(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    if len(result) > limit:
        return "; ".join(result[:limit]) + f"; +{len(result) - limit} more"
    return "; ".join(result)


def parse_year(value):
    if isinstance(value, int) and 1900 <= value <= 2100:
        return value
    if isinstance(value, float) and value.is_integer() and 1900 <= int(value) <= 2100:
        return int(value)
    text = norm(value)
    if re.fullmatch(r"(19|20)\d{2}", text):
        return int(text)
    return None


def periodish(value):
    text = norm(value).lower()
    return bool(parse_year(value) or re.fullmatch(r"(19|20)\d{2}[/\-](\d{2}|(19|20)\d{2})", text))


def nonempty(value):
    return norm(value) != ""


def row_values(ws):
    max_row = min(ws.max_row or 1, MAX_SCAN_ROWS)
    max_col = min(ws.max_column or 1, MAX_SCAN_COLS)
    return [list(row) for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True)]


def row_nonempty(row):
    return sum(1 for value in row if nonempty(value))


def find_header_row(rows):
    best = None
    best_score = -1
    known = ["year", "date", "geography", "area", "local authority", "region", "measure", "value", "rate", "percent", "number", "total", "sex", "age", "sector", "industry"]
    for idx, row in enumerate(rows[:35]):
        texts = [norm(v).lower() for v in row if nonempty(v)]
        if not texts:
            continue
        blob = " | ".join(texts)
        text_count = sum(1 for text in texts if re.search(r"[a-zA-Z]", text))
        numeric_count = sum(1 for value in row if isinstance(value, (int, float)) and not isinstance(value, bool))
        score = len(texts) + text_count - numeric_count + 3 * sum(1 for word in known if word in blob)
        if score > best_score:
            best = idx
            best_score = score
    return best


def infer_headers(rows):
    idx = find_header_row(rows)
    if idx is None:
        return None, []
    headers = [norm(value) for value in rows[idx]]
    last = 0
    for col, header in enumerate(headers):
        if header or any(col < len(row) and nonempty(row[col]) for row in rows[idx + 1: min(len(rows), idx + 12)]):
            last = col
    return idx, headers[:last + 1]


def find_year_header(rows):
    best = None
    for idx, row in enumerate(rows[:25]):
        year_cols = [(col, parse_year(value)) for col, value in enumerate(row)]
        year_cols = [(col, year) for col, year in year_cols if year]
        if len(year_cols) >= 2 and (best is None or len(year_cols) > len(best[1])):
            best = (idx, year_cols)
    return best


def blob(headers, sheet_name, workbook_name):
    return " ".join([sheet_name, workbook_name] + [h for h in headers if h]).lower()


def detect_from_patterns(headers, sheet_name, workbook_name, patterns):
    text = blob(headers, sheet_name, workbook_name)
    return [label for label, words in patterns.items() if any(word in text for word in words)]


def detect_topic(mapping, workbook_name):
    category = mapping.get("category") or ""
    text = f"{category} {mapping.get('title','')} {workbook_name}".lower()
    found = list(CATEGORY_TOPIC_HINTS.get(category, []))
    for topic in TOPIC_ORDER:
        if topic not in found and any(keyword_matches(text, keyword) for keyword in TOPICS[topic]):
            found.append(topic)
    if category == "poverty-and-deprivation":
        found = [topic for topic in found if topic != "housing"] or ["poverty/deprivation"]
    return found or [mapping.get("category") or "uncategorised"]


def keyword_matches(text, keyword):
    if keyword == "commut":
        return keyword in text
    if re.search(r"[^a-z0-9 ]", keyword) or " " in keyword:
        return keyword in text
    return bool(re.search(rf"\b{re.escape(keyword)}\b", text))


def norm_sheet_name(name):
    return re.sub(r"\s+", " ", norm(name).lower().replace("_", " ").replace("-", " ")).strip()


def is_analysis_sheet_name(name):
    return "analysis" in norm_sheet_name(name)


def is_metadata_sheet_name(name):
    normalized = norm_sheet_name(name)
    return normalized == "meta" or "metadata" in normalized


def is_explicit_raw_sheet_name(name):
    return bool(re.search(r"\braw\b", norm_sheet_name(name)))


def is_excluded_analysis_candidate_sheet_name(name):
    normalized = norm_sheet_name(name)
    return any(
        excluded != "analysis" and (normalized == excluded or excluded in normalized)
        for excluded in SOURCE_DATA_EXCLUDED_SHEET_NAMES
    )


def is_excluded_source_data_sheet_name(name):
    normalized = norm_sheet_name(name)
    return any(normalized == excluded or excluded in normalized for excluded in SOURCE_DATA_EXCLUDED_SHEET_NAMES)


def is_candidate_source_data_sheet_name(name):
    normalized = norm_sheet_name(name)
    return any(normalized == candidate or candidate in normalized for candidate in SOURCE_DATA_SHEET_NAMES)


def is_detailed_source_data_info(info):
    structure = info["raw_structure"]
    if structure.startswith("wide_year_columns"):
        return True
    if structure.startswith("wide_period_columns"):
        return True
    if structure.startswith("long_"):
        return True
    if structure.startswith("tabular_measure_columns"):
        return True
    if structure.startswith("measure_value_rows"):
        return True
    if structure.startswith("tabular_unknown_structure"):
        return bool(info["measure_types"] or info["breakdowns"] or info["counts_available"])
    return False


def detect_raw_structure(rows, headers, sheet_name, workbook_name):
    year_header = find_year_header(rows)
    lower_headers = [h.lower() for h in headers if h]
    nonempty_rows = sum(1 for row in rows if row_nonempty(row) > 0)
    if nonempty_rows == 0:
        return "empty_or_unreadable"
    if year_header:
        return f"wide_year_columns ({len(year_header[1])} detected current-parser year columns)"
    has_time = any(any(term in h for term in ["year", "date", "period", "quarter", "month"]) for h in lower_headers)
    has_value = any(any(term in h for term in ["value", "rate", "percent", "percentage", "number", "count", "total", "proportion", "estimate"]) for h in lower_headers)
    has_geo = any(any(term in h for term in ["geography", "area", "local authority", "region", "district", "la ", "la name", "place"]) for h in lower_headers)
    has_measure = any(any(term in h for term in ["measure", "variable", "indicator", "metric", "category"]) for h in lower_headers)
    periodish_headers = sum(1 for h in headers if periodish(h))
    if periodish_headers >= 2:
        return f"wide_period_columns ({periodish_headers} period columns, not all plain years)"
    if has_time and has_value and has_geo:
        if has_measure:
            return "long_tidy_rows (geography/time/measure/value columns)"
        return "long_or_tabular_rows (geography/time/value columns)"
    if has_geo and has_value:
        return "tabular_measure_columns (geography plus measure/value columns)"
    if has_value and has_measure:
        return "measure_value_rows (measure/value columns, geography unclear)"
    if len(lower_headers) >= 3:
        return "tabular_unknown_structure"
    return "unclassified_raw_structure"


def current_parser_support(raw_structure):
    if raw_structure.startswith("wide_year_columns"):
        return "yes - structured wide-year facts should ingest"
    if raw_structure in {"empty_or_unreadable", "unclassified_raw_structure"}:
        return "no - detailed/source-data sheet not parseable from sampled structure"
    return "fallback - searchable source-row snippets should ingest; targeted parser needed for structured calculations"


def has_any(headers, words):
    text = " | ".join(h.lower() for h in headers if h)
    return any(word in text for word in words)


def inspect_sheet(ws, sheet_name, workbook_name):
    rows = row_values(ws)
    header_idx, headers = infer_headers(rows)
    year_header = find_year_header(rows)
    return {
        "sheet_name": sheet_name,
        "header_row": (header_idx + 1) if header_idx is not None else "",
        "headers": headers,
        "raw_structure": detect_raw_structure(rows, headers, sheet_name, workbook_name),
        "plain_year_columns_detected": len(year_header[1]) if year_header else 0,
        "measure_types": detect_from_patterns(headers, sheet_name, workbook_name, MEASURE_PATTERNS),
        "breakdowns": detect_from_patterns(headers, sheet_name, workbook_name, BREAKDOWN_PATTERNS),
        "counts_available": has_any(headers, ["count", "number", "cohort", "population", "total", "n_people"]),
        "numerator_denominator_available": has_any(headers, ["numerator", "denominator", "cohort", "base", "number neet", "number activity"]),
        "confidence_intervals_available": has_any(headers, ["confidence", "ci", "lower", "upper", "lci", "uci", "interval"]),
    }


def find_folder_entries(zip_file, expected_folder):
    expected = norm_path(expected_folder)
    norm_pairs = [(norm_path(name), name) for name in zip_file.namelist()]
    prefix = expected.rstrip("/") + "/"
    direct = [orig for normed, orig in norm_pairs if normed.startswith(prefix)]
    if direct:
        return direct
    tail = "/".join(expected.split("/")[-3:])
    fallback = [orig for normed, orig in norm_pairs if tail in normed]
    if fallback:
        return fallback
    tail_one = expected.split("/")[-1]
    parts = expected.split("/")
    category = parts[2] if len(parts) > 2 else ""
    return [orig for normed, orig in norm_pairs if tail_one in normed and (not category or f"/{category}/" in f"/{normed}/")]


def inspect_workbook(zip_file, zip_name, mapping):
    wb = load_workbook(io.BytesIO(zip_file.read(zip_name)), read_only=True, data_only=True)
    workbook_name = os.path.basename(zip_name)
    sheet_names = list(wb.sheetnames)
    sheet_infos = {sheet_name: inspect_sheet(wb[sheet_name], sheet_name, workbook_name) for sheet_name in sheet_names}
    analysis_sheets = [s for s in sheet_names if is_analysis_sheet_name(s)]
    raw_sheets = [s for s in sheet_names if is_explicit_raw_sheet_name(s) and not is_excluded_source_data_sheet_name(s)]
    metadata_sheets = [s for s in sheet_names if is_metadata_sheet_name(s)]
    excluded_non_data_sheets = [
        s for s in sheet_names
        if is_excluded_source_data_sheet_name(s) and not is_metadata_sheet_name(s) and not is_analysis_sheet_name(s)
    ]
    chosen_analysis = analysis_sheets[0] if analysis_sheets else next(
        (
            s for s in sheet_names
            if not is_metadata_sheet_name(s)
            and not is_explicit_raw_sheet_name(s)
            and not is_excluded_analysis_candidate_sheet_name(s)
        ),
        sheet_names[0] if sheet_names else ""
    )
    analysis_info = sheet_infos.get(chosen_analysis)
    candidate_source_data_sheets = [
        s for s in sheet_names
        if s not in raw_sheets
        and not is_excluded_source_data_sheet_name(s)
        and is_candidate_source_data_sheet_name(s)
        and is_detailed_source_data_info(sheet_infos[s])
    ]
    source_data_sheets = raw_sheets + [s for s in candidate_source_data_sheets if s not in raw_sheets]
    source_infos = [sheet_infos[s] for s in source_data_sheets]
    inspected = source_infos + ([analysis_info] if analysis_info else [])
    measure_types = list(dict.fromkeys(mt for info in inspected for mt in info["measure_types"]))
    breakdowns = list(dict.fromkeys(br for info in inspected for br in info["breakdowns"]))
    counts = any(info["counts_available"] for info in inspected)
    numden = any(info["numerator_denominator_available"] for info in inspected)
    ci = any(info["confidence_intervals_available"] for info in inspected)
    source_structures = [info["raw_structure"] for info in source_infos]

    if source_infos:
        statuses = [current_parser_support(info["raw_structure"]) for info in source_infos]
        if all(status.startswith("yes") for status in statuses):
            parser_status = "yes - structured wide-year facts should ingest from detailed/source-data sheets"
        elif all(status.startswith("fallback") for status in statuses):
            parser_status = "fallback - searchable source-row snippets should ingest; targeted parser needed for structured calculations"
        elif any(status.startswith("yes") or status.startswith("fallback") for status in statuses):
            parser_status = "mixed - some detailed/source-data sheets become structured facts or fallback snippets; others need manual/targeted logic"
        else:
            parser_status = "no - analysis rows ingest, but source-data facts likely not produced by current parser"
    else:
        parser_status = "not applicable for raw facts - no detailed/source-data sheet detected; analysis rows still ingest"

    warnings = []
    if not analysis_sheets:
        warnings.append("No sheet name containing 'analysis'; parser will choose first non-excluded sheet.")
    if not source_data_sheets:
        warnings.append("No detailed/source-data sheet detected; deeper calculation must rely on article/analysis rows or another source.")
    for info in source_infos:
        status = current_parser_support(info["raw_structure"])
        if status.startswith("fallback"):
            warnings.append(f"Detailed/source-data sheet '{info['sheet_name']}' is {info['raw_structure']}; it should ingest as searchable row snippets, but targeted parsing is still needed for structured calculations.")
        elif not status.startswith("yes"):
            warnings.append(f"Detailed/source-data sheet '{info['sheet_name']}' is {info['raw_structure']}; current parser may not ingest it into source-data facts.")
    if counts and not numden:
        warnings.append("Counts/totals appear available, but numerator/denominator pairs are not clearly labelled in sampled headers.")
    if not counts and any(mt in measure_types for mt in ["rates/percentages/proportions", "ratios/prices"]):
        warnings.append("Rates/ratios appear available, but sampled headers do not show counts or denominators.")

    key_headers = []
    if analysis_info:
        key_headers.extend(analysis_info["headers"][:12])
    for info in source_infos:
        key_headers.extend(info["headers"][:12])

    return {
        "slug": mapping.get("slug", ""),
        "public_post_title": norm(mapping.get("title", "")),
        "public_post_url": mapping.get("post_url", ""),
        "category": mapping.get("category", ""),
        "topic_family": compact(detect_topic(mapping, workbook_name), 6),
        "github_path": norm(mapping.get("github_path", "")),
        "zip_folder": norm(mapping.get("zip_folder", "")),
        "workbook_path": zip_name,
        "workbook_file_name": workbook_name,
        "sheets_detected": compact(sheet_names, 30),
        "analysis_sheet_exists": "yes" if analysis_sheets else "no",
        "chosen_analysis_sheet": chosen_analysis,
        "raw_sheets_exist": "yes" if raw_sheets else "no",
        "raw_sheet_names": compact(raw_sheets, 20),
        "candidate_source_data_sheet_names": compact(candidate_source_data_sheets, 20),
        "detailed_source_data_sheets_exist": "yes" if source_data_sheets else "no",
        "detailed_source_data_sheet_names": compact(source_data_sheets, 25),
        "excluded_non_data_sheets": compact(excluded_non_data_sheets, 20),
        "metadata_sheets": compact(metadata_sheets, 10),
        "raw_sheet_structure": compact(source_structures, 20) if source_structures else "none detected",
        "source_data_sheet_structure": compact(source_structures, 20) if source_structures else "none detected",
        "key_columns": compact(key_headers, 28),
        "available_measure_types": compact(measure_types, 20),
        "supported_breakdowns": compact(breakdowns, 18),
        "counts_available": "yes" if counts else "no/unclear from sampled headers",
        "numerators_denominators_available": "yes" if numden else "no/unclear from sampled headers",
        "confidence_intervals_available": "yes" if ci else "no/unclear from sampled headers",
        "current_parser_can_ingest": parser_status,
        "limitations_or_warnings": " ".join(warnings) if warnings else "No obvious structural warning from sampled rows.",
        "workbook_count_for_post": "",
        "matched_workbook_count_for_post": "",
        "folder_match_warning": "",
    }


def missing_row(mapping):
    return {
        "slug": mapping.get("slug", ""),
        "public_post_title": norm(mapping.get("title", "")),
        "public_post_url": mapping.get("post_url", ""),
        "category": mapping.get("category", ""),
        "topic_family": compact(detect_topic(mapping, ""), 6),
        "github_path": norm(mapping.get("github_path", "")),
        "zip_folder": norm(mapping.get("zip_folder", "")),
        "workbook_path": "",
        "workbook_file_name": "",
        "sheets_detected": "",
        "analysis_sheet_exists": "no",
        "chosen_analysis_sheet": "",
        "raw_sheets_exist": "no",
        "raw_sheet_names": "",
        "candidate_source_data_sheet_names": "",
        "detailed_source_data_sheets_exist": "no",
        "detailed_source_data_sheet_names": "",
        "excluded_non_data_sheets": "",
        "metadata_sheets": "",
        "raw_sheet_structure": "none detected",
        "source_data_sheet_structure": "none detected",
        "key_columns": "",
        "available_measure_types": "",
        "supported_breakdowns": "",
        "counts_available": "no/unclear from sampled headers",
        "numerators_denominators_available": "no/unclear from sampled headers",
        "confidence_intervals_available": "no/unclear from sampled headers",
        "current_parser_can_ingest": "not inspected - no workbook matched in local ZIP",
        "limitations_or_warnings": "No .xlsx workbook found under mapped folder in the local ZIP.",
        "workbook_count_for_post": 0,
        "matched_workbook_count_for_post": 0,
        "folder_match_warning": "No workbook matched mapping.zip_folder in local ZIP.",
    }


def counter_md(counter, limit=20):
    if not counter:
        return "- None detected\n"
    return "\n".join(f"- {key}: {count}" for key, count in counter.most_common(limit)) + "\n"


def write_markdown(rows, mappings):
    workbook_rows = [row for row in rows if row.get("workbook_file_name")]
    slugs_with_workbooks = {row["slug"] for row in workbook_rows}
    slugs_with_raw = {row["slug"] for row in workbook_rows if row.get("raw_sheets_exist") == "yes"}
    slugs_with_source_data = {row["slug"] for row in workbook_rows if row.get("detailed_source_data_sheets_exist") == "yes"}
    slugs_without_any_source_data = slugs_with_workbooks - slugs_with_source_data
    slugs_with_mixed_source_data = {
        slug for slug in slugs_with_workbooks
        if any(row["slug"] == slug and row.get("detailed_source_data_sheets_exist") == "yes" for row in workbook_rows)
        and any(row["slug"] == slug and row.get("detailed_source_data_sheets_exist") != "yes" for row in workbook_rows)
    }
    raw_structure_counter = Counter()
    parser_counter = Counter()
    category_counter = Counter()
    measure_counter = Counter()
    breakdown_counter = Counter()
    topic_coverage = defaultdict(list)
    for row in workbook_rows:
        category_counter[row.get("category", "")] += 1
        for item in [s.strip() for s in row.get("raw_sheet_structure", "").split(";") if s.strip()]:
            raw_structure_counter[item] += 1
        parser_counter[row.get("current_parser_can_ingest", "")] += 1
        for item in [s.strip() for s in row.get("available_measure_types", "").split(";") if s.strip()]:
            measure_counter[item] += 1
        for item in [s.strip() for s in row.get("supported_breakdowns", "").split(";") if s.strip()]:
            breakdown_counter[item] += 1
        for item in [s.strip() for s in row.get("topic_family", "").split(";") if s.strip()]:
            topic_coverage[item].append(row)

    analysis_yes = sum(1 for row in workbook_rows if row.get("analysis_sheet_exists") == "yes")
    analysis_no = sum(1 for row in workbook_rows if row.get("analysis_sheet_exists") != "yes")

    md = [
        "# Dataset Measure Audit",
        "",
        f"Generated from the local ZIP: `{ZIP_PATH}`",
        "",
        "## Scope",
        f"- Mapped Data Hub posts inspected: {len(mappings)}",
        f"- Mapped posts with matched workbooks in the local ZIP: {len(slugs_with_workbooks)}",
        f"- Workbook records inspected: {len(workbook_rows)}",
        f"- Posts with at least one explicit raw sheet in any matched workbook: {len(slugs_with_raw)}",
        f"- Posts with at least one detailed/source-data sheet in any matched workbook: {len(slugs_with_source_data)}",
        f"- Posts with no detailed/source-data sheets in any matched workbook: {len(slugs_without_any_source_data)}",
        f"- Posts with mixed workbook-level detailed/source-data coverage: {len(slugs_with_mixed_source_data)}",
        f"- Workbooks with an explicit analysis sheet: {analysis_yes}",
        f"- Workbooks without an explicit analysis sheet name: {analysis_no}",
        "",
        "## Important Interpretation",
        "This audit samples workbook headers and early rows to describe structure. It is designed to guide parser and answer-logic decisions, not to certify every individual numeric value. The current chatbot dataset ingestion already stores analysis-sheet rows for lookup-style questions. Detailed/source-data sheets include explicit raw sheets plus candidate sheets such as `Data`, `source_data`, `table`, or `observations` when their headers look like detailed data. Raw facts are still only structured when the current parser recognises plain wide year columns.",
        "",
        "## Detailed/Source-Data Sheet Structures Detected",
        counter_md(raw_structure_counter),
        "## Current Parser Coverage For Detailed/Source-Data Sheets",
        counter_md(parser_counter),
        "## Available Measure Types Detected",
        counter_md(measure_counter),
        "## Supported Breakdowns Detected",
        counter_md(breakdown_counter),
        "## Category Workbook Counts",
        counter_md(category_counter),
        "## Representative Topic Samples",
    ]

    for topic in TOPIC_ORDER:
        matches = sorted(topic_coverage.get(topic, []), key=lambda row: topic_sample_priority(row, topic))
        md.append(f"### {topic}")
        if not matches:
            md.append("No mapped workbook was classified into this topic from title/category keywords.\n")
            continue
        picked = []
        for candidate in [
            next((r for r in matches if r.get("detailed_source_data_sheets_exist") == "yes"), None),
            next((r for r in matches if r.get("detailed_source_data_sheets_exist") != "yes"), None),
        ]:
            if candidate and candidate not in picked:
                picked.append(candidate)
        if not picked:
            picked = matches[:2]
        for item in picked[:3]:
            md.extend([
                f"- **{item['public_post_title']}** | workbook `{item['workbook_file_name']}`",
                f"  - Sheets: {item['sheets_detected'] or 'not detected'}",
                f"  - Analysis: {item['analysis_sheet_exists']} (`{item['chosen_analysis_sheet']}`); Explicit raw: {item['raw_sheets_exist']} ({item['raw_sheet_names'] or 'none'}); Detailed/source-data: {item['detailed_source_data_sheets_exist']} ({item['detailed_source_data_sheet_names'] or 'none'})",
                f"  - Detailed/source-data structure: {item['source_data_sheet_structure']}",
                f"  - Measures: {item['available_measure_types'] or 'unclear from sampled headers'}",
                f"  - Breakdowns: {item['supported_breakdowns'] or 'unclear from sampled headers'}",
                f"  - Parser status: {item['current_parser_can_ingest']}",
            ])
        md.append("")

    unsupported = [
        row for row in workbook_rows
        if row.get("current_parser_can_ingest", "").startswith(("fallback", "mixed", "no"))
    ]
    md.extend(["## Raw Parser Limitations / Warnings"])
    if unsupported:
        for item in unsupported[:18]:
            md.append(f"- **{item['public_post_title']}** / `{item['workbook_file_name']}`: {item['source_data_sheet_structure']} ({item['current_parser_can_ingest']})")
    else:
        md.append("- None detected in sampled workbooks.")
    md.extend([
        "",
        "## Practical Guidance For Chatbot Behaviour",
        "- Keep the current article-first, analysis-second, raw-third hierarchy.",
        "- Do not claim a raw value is unavailable until article context and analysis-sheet rows have also been checked.",
        "- Do not force count logic across all datasets. Many datasets expose rates, percentages, ratios, prices, index values, emissions totals, or wide time-series values without explicit numerator/denominator columns.",
        "- For calculation questions, only calculate where the needed numerator/denominator/count fields are clearly present and matched to the requested geography, year, measure, and breakdown.",
        "- Where detailed/source-data sheets exist but are not plain wide-year layouts, the safest next parser improvement is a fallback searchable source-row snippet store, plus targeted parsers for common long/tidy geography-year-measure-value structures.",
        "",
        "## Detailed Inventory",
        f"See `{CSV_OUT.name}` for the per-workbook audit fields requested: title, URL, workbook, sheets, analysis/raw/source-data availability, source-data structure, key columns, measures, breakdowns, counts, numerator/denominator flags, confidence interval flags, parser status, and warnings.",
        "",
    ])
    MD_OUT.write_text("\n".join(md), encoding="utf-8")


def topic_sample_priority(row, topic):
    preferred = TOPIC_PREFERRED_CATEGORIES.get(topic, [])
    category = row.get("category", "")
    category_score = 0 if category in preferred else 1
    sample_text = f"{row.get('public_post_title', '')} {row.get('workbook_file_name', '')}".lower()
    keyword_score = 0 if any(keyword_matches(sample_text, keyword) for keyword in TOPIC_SAMPLE_KEYWORDS.get(topic, [])) else 1
    source_data_score = 0 if row.get("detailed_source_data_sheets_exist") == "yes" else 1
    return (category_score, keyword_score, source_data_score, row.get("public_post_title", ""), row.get("workbook_file_name", ""))


def main():
    mappings = json.loads(MAP_PATH.read_text(encoding="utf-8-sig"))
    rows = []
    workbook_counts = {}
    with zipfile.ZipFile(ZIP_PATH, "r") as zip_file:
        for mapping in mappings:
            entries = find_folder_entries(zip_file, mapping.get("zip_folder", ""))
            workbooks = sorted(
                name for name in entries
                if name.lower().endswith(".xlsx") and not os.path.basename(name).startswith("~$")
            )
            workbook_counts[mapping.get("slug", "")] = len(workbooks)
            if not workbooks:
                rows.append(missing_row(mapping))
                continue
            for workbook in workbooks:
                try:
                    rows.append(inspect_workbook(zip_file, workbook, mapping))
                except Exception as exc:
                    failed = missing_row(mapping)
                    failed.update({
                        "workbook_path": workbook,
                        "workbook_file_name": os.path.basename(workbook),
                        "raw_sheet_structure": "inspection_failed",
                        "source_data_sheet_structure": "inspection_failed",
                        "current_parser_can_ingest": "unknown - workbook inspection failed",
                        "limitations_or_warnings": f"Workbook inspection failed: {exc}",
                    })
                    rows.append(failed)

    for row in rows:
        count = workbook_counts.get(row["slug"], "")
        row["workbook_count_for_post"] = count
        row["matched_workbook_count_for_post"] = count

    fields = [
        "slug", "public_post_title", "public_post_url", "category", "topic_family", "github_path", "zip_folder",
        "workbook_count_for_post", "matched_workbook_count_for_post", "folder_match_warning",
        "workbook_path", "workbook_file_name", "sheets_detected", "analysis_sheet_exists", "chosen_analysis_sheet",
        "raw_sheets_exist", "raw_sheet_names", "candidate_source_data_sheet_names",
        "detailed_source_data_sheets_exist", "detailed_source_data_sheet_names",
        "excluded_non_data_sheets", "metadata_sheets", "raw_sheet_structure", "source_data_sheet_structure", "key_columns",
        "available_measure_types", "supported_breakdowns", "counts_available", "numerators_denominators_available",
        "confidence_intervals_available", "current_parser_can_ingest", "limitations_or_warnings",
    ]
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})

    write_markdown(rows, mappings)
    workbook_rows = [row for row in rows if row.get("workbook_file_name")]
    slugs_with_workbooks = {row["slug"] for row in workbook_rows}
    slugs_with_raw = {row["slug"] for row in workbook_rows if row.get("raw_sheets_exist") == "yes"}
    slugs_with_source_data = {row["slug"] for row in workbook_rows if row.get("detailed_source_data_sheets_exist") == "yes"}
    slugs_without_any_source_data = slugs_with_workbooks - slugs_with_source_data
    print(json.dumps({
        "mapped_posts": len(mappings),
        "mapped_posts_with_workbooks": len(slugs_with_workbooks),
        "workbook_records": len(workbook_rows),
        "posts_with_at_least_one_raw_sheet": len(slugs_with_raw),
        "posts_with_at_least_one_detailed_source_data_sheet": len(slugs_with_source_data),
        "posts_with_no_detailed_source_data_sheets": len(slugs_without_any_source_data),
        "workbooks_with_analysis_sheet": sum(1 for row in workbook_rows if row.get("analysis_sheet_exists") == "yes"),
        "csv": str(CSV_OUT),
        "markdown": str(MD_OUT),
    }, indent=2))


if __name__ == "__main__":
    main()
